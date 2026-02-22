"""
KVP-Funda Scanner
=================
Vergelijkt automatisch de kamerverhuurvergunninglijst van gemeente Tilburg
met het actuele koopwoningaanbod op Funda.

Gebruik:
    python kvp_scanner.py                  # Eenmalig scannen
    python kvp_scanner.py --schedule       # Dagelijks automatisch scannen
    python kvp_scanner.py --email je@mail.nl  # Met e-mailnotificatie

Vereisten:
    pip install pdfplumber requests beautifulsoup4 openpyxl
"""

import re
import time
import json
import argparse
import smtplib
import unicodedata
import logging
from datetime import datetime
from pathlib import Path
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

import requests
import pdfplumber
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ─── Configuratie ────────────────────────────────────────────────────────────

KVP_PDF_URL = "https://www.tilburg.nl/fileadmin/files/inwoners/vergunningen/lijst_kvp_internet_251212.pdf"
FUNDA_BASE_URL = "https://www.funda.nl/koop/tilburg/straat-{slug}/"
OUTPUT_DIR = Path("output")
RESULTS_JSON = OUTPUT_DIR / "resultaten.json"
RESULTS_EXCEL = OUTPUT_DIR / "kvp_funda_matches.xlsx"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "nl-NL,nl;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

# Vertraging tussen Funda-verzoeken (seconden) — wees beleefd voor de server
REQUEST_DELAY = 2.0

# ─── Logging ─────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(OUTPUT_DIR / "scanner.log" if OUTPUT_DIR.exists() else "scanner.log"),
    ],
)
log = logging.getLogger(__name__)


# ─── Stap 1: PDF inlezen ─────────────────────────────────────────────────────

def slugify(text: str) -> str:
    """Zet een straatnaam om naar een Funda URL-slug."""
    # Verwijder diacritics (é → e, ë → e, etc.)
    normalized = unicodedata.normalize("NFKD", text)
    ascii_text = normalized.encode("ascii", "ignore").decode("ascii")
    # Verwijder speciale tekens, zet spaties om naar koppeltekens
    slug = re.sub(r"[^\w\s-]", "", ascii_text).strip().lower()
    slug = re.sub(r"[\s_]+", "-", slug)
    return slug


def parse_address_line(line: str) -> tuple[str, str, str] | None:
    """
    Parseer een regel als 'Straatnaam 123 5011AB' of 'Straatnaam 12 A 5011AB'.
    Geeft (straatnaam, huisnummer, postcode) terug of None.
    """
    line = line.strip()
    # Patroon: alles t/m het eerste cijfer = straatnaam, dan huisnummer, dan postcode
    match = re.match(
        r"^(.+?)\s+(\d+\s*[A-Za-z0-9\s\-]*?)\s+(\d{4}\s*[A-Z]{2})\s*$",
        line,
        re.IGNORECASE,
    )
    if not match:
        return None
    straat = match.group(1).strip()
    nummer = match.group(2).strip()
    postcode = match.group(3).replace(" ", "").upper()
    # Filter koptekstregels
    if straat.lower() in {"straatnaam", "kamerverhuur-", "huisnummer"}:
        return None
    return straat, nummer, postcode


def download_and_parse_kvp_pdf(url: str) -> list[dict]:
    """Download de KVP-PDF en parseer alle adressen."""
    log.info("PDF downloaden van %s", url)
    resp = requests.get(url, headers=HEADERS, timeout=30)
    resp.raise_for_status()

    pdf_path = OUTPUT_DIR / "kvp_lijst.pdf"
    pdf_path.write_bytes(resp.content)
    log.info("PDF opgeslagen: %s", pdf_path)

    addresses = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            for line in text.splitlines():
                result = parse_address_line(line)
                if result:
                    straat, nummer, postcode = result
                    addresses.append({
                        "straatnaam": straat,
                        "huisnummer": nummer,
                        "postcode": postcode,
                        "volledig_adres": f"{straat} {nummer}, {postcode} Tilburg",
                    })

    log.info("✓ %d adressen ingelezen uit de KVP-lijst", len(addresses))
    return addresses


# ─── Stap 2: Funda per straat checken ────────────────────────────────────────

def get_funda_listings_for_street(street_name: str) -> list[dict]:
    """
    Vraag alle actieve koopwoningen op voor één straat op Funda.
    Geeft een lijst van dicts terug met adres, prijs, URL, etc.
    """
    slug = slugify(street_name)
    url = FUNDA_BASE_URL.format(slug=slug)

    try:
        resp = requests.get(url, headers=HEADERS, timeout=15)
        if resp.status_code == 404:
            return []
        resp.raise_for_status()
    except requests.RequestException as e:
        log.warning("Fout bij ophalen %s: %s", url, e)
        return []

    soup = BeautifulSoup(resp.text, "html.parser")
    listings = []

    # Zoek alle woningkaarten op de pagina
    for card in soup.find_all("a", href=re.compile(r"/detail/koop/tilburg/")):
        href = card.get("href", "")
        if not href:
            continue

        # Haal adres uit de link of omliggende tekst
        full_url = f"https://www.funda.nl{href}" if href.startswith("/") else href

        # Zoek huisnummer in de URL (bv. /huis-korvelseweg-81/)
        addr_match = re.search(r"/(?:huis|appartement)-(.+?)/(\d+[a-z0-9\-]*)/", href)
        if not addr_match:
            continue

        street_slug_found = addr_match.group(1)
        huisnummer_raw = addr_match.group(2)

        # Zoek prijsinfo in de omliggende context
        prijs_el = card.find_parent("li") or card.find_parent("div", class_=re.compile(r"listing|result|property", re.I))
        prijs_text = ""
        status = "Te koop"

        if prijs_el:
            prijs_match = re.search(r"€[\s\d\.,]+", prijs_el.get_text())
            if prijs_match:
                prijs_text = prijs_match.group(0).strip()
            if "onder voorbehoud" in prijs_el.get_text().lower():
                status = "Onder voorbehoud"
            if "verkocht" in prijs_el.get_text().lower() and "onder voorbehoud" not in prijs_el.get_text().lower():
                status = "Verkocht"

        listings.append({
            "funda_url": full_url,
            "straatnaam_funda": street_name,
            "huisnummer_funda": huisnummer_raw,
            "prijs": prijs_text,
            "status": status,
        })

    # Verwijder duplicaten op basis van URL
    seen = set()
    unique = []
    for l in listings:
        if l["funda_url"] not in seen:
            seen.add(l["funda_url"])
            unique.append(l)

    return unique


def normalize_number(num: str) -> str:
    """Normaliseer huisnummer voor vergelijking (lowercase, geen spaties)."""
    return re.sub(r"\s+", "", num).lower()


def numbers_match(kvp_num: str, funda_num: str) -> bool:
    """Controleer of twee huisnummers overeenkomen (met tolerantie voor toevoegingen)."""
    k = normalize_number(kvp_num)
    f = normalize_number(funda_num)
    # Exacte match
    if k == f:
        return True
    # KVP heeft '29', Funda heeft '29-a' of '29a'
    if f.startswith(k) and not f[len(k):len(k)+1].isdigit():
        return True
    # KVP heeft '29 A', Funda heeft '29'
    if k.startswith(f) and not k[len(f):len(f)+1].isdigit():
        return True
    return False


# ─── Stap 3: Vergelijken ─────────────────────────────────────────────────────

def find_matches(kvp_addresses: list[dict]) -> list[dict]:
    """
    Vergelijk de KVP-lijst met Funda.
    Geeft alle matches terug (zowel directe als straat-level).
    """
    # Groepeer KVP-adressen per straat
    kvp_by_street: dict[str, list[dict]] = {}
    for addr in kvp_addresses:
        straat = addr["straatnaam"]
        kvp_by_street.setdefault(straat, []).append(addr)

    all_streets = sorted(kvp_by_street.keys())
    log.info("Funda checken voor %d unieke straten...", len(all_streets))

    matches = []
    scan_log = []

    for i, straat in enumerate(all_streets, 1):
        if i % 50 == 0:
            log.info("  Voortgang: %d/%d straten gescand", i, len(all_streets))

        funda_listings = get_funda_listings_for_street(straat)

        for listing in funda_listings:
            # Vergelijk elk Funda-resultaat met KVP-adressen op die straat
            for kvp_addr in kvp_by_street[straat]:
                if numbers_match(kvp_addr["huisnummer"], listing["huisnummer_funda"]):
                    match = {
                        **kvp_addr,
                        **listing,
                        "gevonden_op": datetime.now().isoformat(),
                        "match_type": "exact" if normalize_number(kvp_addr["huisnummer"]) == normalize_number(listing["huisnummer_funda"]) else "gedeeltelijk",
                    }
                    matches.append(match)
                    log.info(
                        "  ✅ MATCH: %s %s — %s (%s)",
                        straat, kvp_addr["huisnummer"],
                        listing.get("prijs", "prijs onbekend"),
                        listing["status"],
                    )

        scan_log.append({
            "straat": straat,
            "funda_resultaten": len(funda_listings),
            "matches": sum(1 for m in matches if m["straatnaam"] == straat),
        })

        time.sleep(REQUEST_DELAY)

    log.info("─" * 60)
    log.info("Scan voltooid: %d matches gevonden uit %d straten", len(matches), len(all_streets))
    return matches, scan_log


# ─── Stap 4: Excel exporteren ─────────────────────────────────────────────────

ORANGE = "E87722"
DARK = "1A1A2E"
GREEN = "2ECC71"
AMBER = "F39C12"
LIGHT_GRAY = "F5F5F5"
MID_GRAY = "CCCCCC"

def _header_style(ws, row, cols):
    fill = PatternFill("solid", start_color=DARK, end_color=DARK)
    font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def export_excel(matches: list[dict], scan_log: list[dict]) -> Path:
    OUTPUT_DIR.mkdir(exist_ok=True)
    wb = Workbook()

    # ── Tabblad 1: Matches ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "✅ Matches"
    ws1.row_dimensions[1].height = 30

    headers = [
        "Straatnaam (KVP)", "Huisnummer (KVP)", "Postcode",
        "Volledig adres", "Status Funda", "Vraagprijs",
        "Funda URL", "Match type", "Gevonden op"
    ]
    ws1.append(headers)
    _header_style(ws1, 1, len(headers))

    status_colors = {
        "Te koop": "D5F5E3",
        "Onder voorbehoud": "FDEBD0",
        "Verkocht": "FADBD8",
    }

    for i, m in enumerate(matches, 2):
        row = [
            m.get("straatnaam", ""),
            m.get("huisnummer", ""),
            m.get("postcode", ""),
            m.get("volledig_adres", ""),
            m.get("status", ""),
            m.get("prijs", ""),
            m.get("funda_url", ""),
            m.get("match_type", ""),
            m.get("gevonden_op", "")[:10] if m.get("gevonden_op") else "",
        ]
        ws1.append(row)

        color = status_colors.get(m.get("status", ""), "FFFFFF")
        fill = PatternFill("solid", start_color=color, end_color=color)
        for col in range(1, len(headers) + 1):
            cell = ws1.cell(row=i, column=col)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            cell.font = Font(name="Arial", size=10)

        # URL als hyperlink
        url_cell = ws1.cell(row=i, column=7)
        if url_cell.value:
            url_cell.hyperlink = url_cell.value
            url_cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")
            url_cell.value = "Bekijk op Funda →"

    col_widths = [25, 15, 12, 40, 18, 16, 22, 15, 14]
    for col, width in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(col)].width = width

    # Overzicht onderaan
    if matches:
        ws1.append([])
        summary_row = ws1.max_row + 1
        ws1.cell(row=summary_row, column=1, value="Totaal matches:")
        ws1.cell(row=summary_row, column=2, value=f'=COUNTA(B2:B{summary_row - 2})')
        ws1.cell(row=summary_row, column=1).font = Font(bold=True, name="Arial")

    # ── Tabblad 2: Alle gescande straten ────────────────────────────────────
    ws2 = wb.create_sheet("📋 Scan log")
    ws2.row_dimensions[1].height = 28
    ws2.append(["Straat", "Funda resultaten", "KVP matches"])
    _header_style(ws2, 1, 3)

    for entry in scan_log:
        row_data = [entry["straat"], entry["funda_resultaten"], entry["matches"]]
        ws2.append(row_data)
        r = ws2.max_row
        for col in range(1, 4):
            cell = ws2.cell(row=r, column=col)
            cell.font = Font(name="Arial", size=10)
            if entry["matches"] > 0:
                cell.fill = PatternFill("solid", start_color="D5F5E3", end_color="D5F5E3")

    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 20
    ws2.column_dimensions["C"].width = 15

    # ── Tabblad 3: Samenvatting ──────────────────────────────────────────────
    ws3 = wb.create_sheet("📊 Samenvatting")

    summary = [
        ("Scan datum", datetime.now().strftime("%d-%m-%Y %H:%M")),
        ("Bron KVP-lijst", KVP_PDF_URL),
        ("Totaal KVP-adressen", ""),
        ("Unieke straten gescand", len(scan_log)),
        ("Totaal matches gevonden", len(matches)),
        ("Waarvan 'Te koop'", sum(1 for m in matches if m.get("status") == "Te koop")),
        ("Waarvan 'Onder voorbehoud'", sum(1 for m in matches if m.get("status") == "Onder voorbehoud")),
        ("Waarvan 'Verkocht'", sum(1 for m in matches if m.get("status") == "Verkocht")),
    ]

    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 55

    for r, (label, value) in enumerate(summary, 2):
        ws3.cell(row=r, column=1, value=label).font = Font(bold=True, name="Arial", size=11)
        ws3.cell(row=r, column=2, value=value).font = Font(name="Arial", size=11)
        ws3.row_dimensions[r].height = 22

    fill_orange = PatternFill("solid", start_color=ORANGE, end_color=ORANGE)
    ws3["A1"] = "KVP × Funda Scanner — Samenvatting"
    ws3["A1"].font = Font(bold=True, name="Arial", size=16, color="FFFFFF")
    ws3["A1"].fill = fill_orange
    ws3.merge_cells("A1:B1")
    ws3.row_dimensions[1].height = 36

    wb.save(RESULTS_EXCEL)
    log.info("Excel opgeslagen: %s", RESULTS_EXCEL)
    return RESULTS_EXCEL


# ─── Stap 5: JSON opslaan (voor dashboard) ───────────────────────────────────

def save_json(matches: list[dict], scan_log: list[dict], kvp_count: int):
    OUTPUT_DIR.mkdir(exist_ok=True)
    payload = {
        "scan_datum": datetime.now().isoformat(),
        "kvp_adressen_totaal": kvp_count,
        "straten_gescand": len(scan_log),
        "matches": matches,
        "scan_log": scan_log,
    }
    RESULTS_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2))
    log.info("JSON opgeslagen: %s", RESULTS_JSON)


# ─── Stap 6: E-mail verzenden ─────────────────────────────────────────────────

def send_email_notification(
    matches: list[dict],
    recipient: str,
    smtp_host: str = "smtp.gmail.com",
    smtp_port: int = 587,
    smtp_user: str = "",
    smtp_pass: str = "",
):
    """
    Stuur een e-mail met een overzicht van alle matches.

    Tip: gebruik een Gmail App Password via:
    Google Account → Beveiliging → App-wachtwoorden
    """
    if not matches:
        log.info("Geen matches — geen e-mail verzonden")
        return

    msg = MIMEMultipart("alternative")
    msg["Subject"] = f"🏠 KVP×Funda: {len(matches)} match(es) gevonden — {datetime.now().strftime('%d-%m-%Y')}"
    msg["From"] = smtp_user
    msg["To"] = recipient

    # Plain text versie
    plain_lines = [f"KVP×Funda Scanner — {datetime.now().strftime('%d-%m-%Y %H:%M')}", "=" * 50, ""]
    for m in matches:
        plain_lines.append(f"• {m.get('volledig_adres', '')} — {m.get('prijs', '')} ({m.get('status', '')})")
        plain_lines.append(f"  {m.get('funda_url', '')}")
        plain_lines.append("")
    msg.attach(MIMEText("\n".join(plain_lines), "plain"))

    # HTML versie
    rows_html = ""
    for m in matches:
        color = {"Te koop": "#d5f5e3", "Onder voorbehoud": "#fdebd0", "Verkocht": "#fadbd8"}.get(m.get("status", ""), "#fff")
        rows_html += f"""
        <tr style="background:{color}">
          <td style="padding:8px;border:1px solid #ddd">{m.get('volledig_adres','')}</td>
          <td style="padding:8px;border:1px solid #ddd"><strong>{m.get('prijs','')}</strong></td>
          <td style="padding:8px;border:1px solid #ddd">{m.get('status','')}</td>
          <td style="padding:8px;border:1px solid #ddd">
            <a href="{m.get('funda_url','')}">Funda →</a>
          </td>
        </tr>"""

    html = f"""
    <html><body style="font-family:Arial,sans-serif;color:#333">
      <h2 style="background:#E87722;color:#fff;padding:16px;border-radius:8px">
        🏠 KVP×Funda — {len(matches)} match(es) gevonden
      </h2>
      <p>Scan van {datetime.now().strftime('%d-%m-%Y om %H:%M')}</p>
      <table style="border-collapse:collapse;width:100%">
        <tr style="background:#1A1A2E;color:#fff">
          <th style="padding:10px;text-align:left">Adres</th>
          <th style="padding:10px;text-align:left">Prijs</th>
          <th style="padding:10px;text-align:left">Status</th>
          <th style="padding:10px;text-align:left">Link</th>
        </tr>
        {rows_html}
      </table>
      <p style="color:#999;font-size:12px;margin-top:24px">
        Automatisch verzonden door KVP×Funda Scanner
      </p>
    </body></html>"""
    msg.attach(MIMEText(html, "html"))

    # Voeg Excel bij als bijlage
    if RESULTS_EXCEL.exists():
        with open(RESULTS_EXCEL, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", f'attachment; filename="{RESULTS_EXCEL.name}"')
            msg.attach(part)

    try:
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.sendmail(smtp_user, recipient, msg.as_string())
        log.info("✉️  E-mail verzonden naar %s", recipient)
    except Exception as e:
        log.error("E-mail versturen mislukt: %s", e)
        log.error("Tip: gebruik een Gmail App Password, niet je normale wachtwoord.")


# ─── Hoofdfunctie ─────────────────────────────────────────────────────────────

def run_scan(email_recipient: str = "", smtp_user: str = "", smtp_pass: str = ""):
    OUTPUT_DIR.mkdir(exist_ok=True)

    log.info("=" * 60)
    log.info("KVP × Funda Scanner gestart — %s", datetime.now().strftime("%d-%m-%Y %H:%M"))
    log.info("=" * 60)

    # 1. PDF inlezen
    kvp_addresses = download_and_parse_kvp_pdf(KVP_PDF_URL)

    # 2. Funda checken per straat + matches vinden
    matches, scan_log = find_matches(kvp_addresses)

    # 3. Resultaten opslaan
    save_json(matches, scan_log, len(kvp_addresses))
    export_excel(matches, scan_log)

    # 4. E-mail sturen (optioneel)
    if email_recipient and smtp_user and smtp_pass:
        send_email_notification(matches, email_recipient, smtp_user=smtp_user, smtp_pass=smtp_pass)

    log.info("=" * 60)
    log.info("Klaar! Resultaten in: %s/", OUTPUT_DIR)
    log.info("  • Excel:   %s", RESULTS_EXCEL)
    log.info("  • JSON:    %s", RESULTS_JSON)
    log.info("  • Log:     %s", OUTPUT_DIR / "scanner.log")
    log.info("=" * 60)

    return matches


# ─── CLI ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="KVP × Funda Scanner")
    parser.add_argument("--email", help="E-mailadres voor notificaties", default="")
    parser.add_argument("--smtp-user", help="Gmail-adres (afzender)", default="")
    parser.add_argument("--smtp-pass", help="Gmail App Password", default="")
    parser.add_argument(
        "--schedule",
        action="store_true",
        help="Dagelijks automatisch scannen om 07:00"
    )
    args = parser.parse_args()

    if args.schedule:
        # Eenvoudige scheduler zonder externe dependency
        import sched, time as time_mod
        log.info("Scheduler actief — scan elke 24 uur om 07:00")
        while True:
            run_scan(args.email, args.smtp_user, args.smtp_pass)
            now = datetime.now()
            next_run = now.replace(hour=7, minute=0, second=0, microsecond=0)
            if next_run <= now:
                from datetime import timedelta
                next_run += timedelta(days=1)
            wait_seconds = (next_run - now).total_seconds()
            log.info("Volgende scan om %s (over %.0f uur)", next_run.strftime("%d-%m %H:%M"), wait_seconds / 3600)
            time_mod.sleep(wait_seconds)
    else:
        run_scan(args.email, args.smtp_user, args.smtp_pass)
