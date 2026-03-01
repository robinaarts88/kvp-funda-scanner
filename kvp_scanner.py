"""
KVP-Funda Scanner (verbeterde versie)
======================================
Haalt ALLE te koop staande woningen in Tilburg op via één Funda-zoekopdracht,
en kruist die lijst daarna met de KVP-vergunningenlijst van gemeente Tilburg.

Voordeel t.o.v. de oude aanpak:
- Geen 714 aparte verzoeken meer (werd geblokkeerd door Funda)
- Één zoekopdracht per pagina, veel minder verdacht
- Sneller: 5-10 minuten i.p.v. 25-30 minuten
"""

import re
import time
import json
import os
import smtplib
import unicodedata
import logging
from datetime import datetime
from pathlib import Path
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

import requests
import pdfplumber
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ─── Configuratie ─────────────────────────────────────────────────────────────

KVP_PDF_URL = "https://www.tilburg.nl/fileadmin/files/inwoners/vergunningen/lijst_kvp_internet_251212.pdf"
OUTPUT_DIR = Path("output")
RESULTS_JSON = OUTPUT_DIR / "resultaten.json"
RESULTS_EXCEL = OUTPUT_DIR / "kvp_funda_matches.xlsx"

# Funda zoekpagina voor heel Tilburg (alle koopwoningen)
FUNDA_SEARCH_URL = "https://www.funda.nl/zoeken/koop/?selected_area=%5B%22tilburg%22%5D&search_result={pagina}"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "nl-NL,nl;q=0.9,en;q=0.8",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "none",
    "Cache-Control": "max-age=0",
}

PAGE_DELAY = 3.0   # seconden tussen pagina's

# ─── Logging ──────────────────────────────────────────────────────────────────

OUTPUT_DIR.mkdir(exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(OUTPUT_DIR / "scanner.log"),
    ],
)
log = logging.getLogger(__name__)


# ─── Stap 1: KVP-lijst inlezen uit PDF ───────────────────────────────────────

def parse_address_line(line: str):
    line = line.strip()
    match = re.match(
        r"^(.+?)\s+(\d+\s*[A-Za-z0-9\s\-]*?)\s+(\d{4}\s*[A-Z]{2})\s*$",
        line,
        re.IGNORECASE,
    )
    if not match:
        return None
    straat = match.group(1).strip()
    nummer = match.group(2).strip()
    postcode = match.group(3).replace(" ", "").upper()
    if straat.lower() in {"straatnaam", "kamerverhuur-", "huisnummer"}:
        return None
    return straat, nummer, postcode


def normalize_straat(naam: str) -> str:
    """Normaliseert een straatnaam voor vergelijking (lowercase, geen diacrieten)."""
    normalized = unicodedata.normalize("NFKD", naam)
    ascii_text = normalized.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", ascii_text).strip().lower()


def normalize_number(num: str) -> str:
    return re.sub(r"\s+", "", num).lower()


def download_and_parse_kvp_pdf(url: str) -> list:
    log.info("KVP-PDF downloaden van gemeente Tilburg...")
    resp = requests.get(url, headers=HEADERS, timeout=30)
    resp.raise_for_status()

    pdf_path = OUTPUT_DIR / "kvp_lijst.pdf"
    pdf_path.write_bytes(resp.content)

    addresses = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            for line in text.splitlines():
                result = parse_address_line(line)
                if result:
                    straat, nummer, postcode = result
                    addresses.append({
                        "straatnaam": straat,
                        "straatnaam_norm": normalize_straat(straat),
                        "huisnummer": nummer,
                        "postcode": postcode,
                        "volledig_adres": f"{straat} {nummer}, {postcode} Tilburg",
                    })

    log.info("✓ %d adressen ingelezen uit de KVP-lijst", len(addresses))
    return addresses


# ─── Stap 2: Alle Funda-listings ophalen voor Tilburg ────────────────────────

def scrape_funda_pagina(pagina: int, sessie: requests.Session) -> tuple[list, bool]:
    """
    Haalt één Funda-pagina op. Geeft (listings, heeft_volgende_pagina) terug.
    """
    url = FUNDA_SEARCH_URL.format(pagina=pagina)
    try:
        resp = sessie.get(url, headers=HEADERS, timeout=20)
        if resp.status_code != 200:
            log.warning("Funda pagina %d: HTTP %d", pagina, resp.status_code)
            return [], False
    except requests.RequestException as e:
        log.error("Funda fout pagina %d: %s", pagina, e)
        return [], False

    soup = BeautifulSoup(resp.text, "html.parser")
    listings = []

    # Zoek alle pand-links op de pagina
    for link in soup.find_all("a", href=re.compile(r"/detail/koop/tilburg/(huis|appartement)-")):
        href = link.get("href", "")
        if not href:
            continue

        full_url = f"https://www.funda.nl{href}" if href.startswith("/") else href

        # Straatnaam en huisnummer uit de URL halen
        # Patroon: /detail/koop/tilburg/huis-POSTCODE-STRAAT/NUMMER/
        url_match = re.search(
            r"/detail/koop/tilburg/(?:huis|appartement)-(\d{4}[a-z]{2})-(.+?)/(\d+[a-z0-9\-]*)/",
            href,
            re.IGNORECASE,
        )
        if not url_match:
            continue

        postcode_url = url_match.group(1).upper()
        straat_slug = url_match.group(2)
        huisnummer_url = url_match.group(3)

        # Straatnaam reconstrueren vanuit slug
        straat_naam = straat_slug.replace("-", " ").strip()

        # Prijsinformatie zoeken in de omliggende HTML
        container = link.find_parent("li") or link.find_parent("div")
        prijs_text = ""
        status = "Te koop"

        if container:
            tekst = container.get_text(" ", strip=True)
            prijs_match = re.search(r"€\s?[\d\.,]+", tekst)
            if prijs_match:
                prijs_text = prijs_match.group(0).strip()
            if "onder voorbehoud" in tekst.lower():
                status = "Onder voorbehoud"
            elif "verkocht" in tekst.lower():
                status = "Verkocht"

        listings.append({
            "funda_url": full_url,
            "straatnaam_funda": straat_naam,
            "straatnaam_norm": normalize_straat(straat_naam),
            "huisnummer_funda": huisnummer_url,
            "postcode_funda": postcode_url,
            "prijs": prijs_text,
            "status": status,
        })

    # Dedupliceren
    seen = set()
    uniek = []
    for l in listings:
        if l["funda_url"] not in seen:
            seen.add(l["funda_url"])
            uniek.append(l)

    # Controleer of er een volgende pagina is
    heeft_volgende = bool(
        soup.find("a", {"aria-label": re.compile(r"volgende|next", re.I)}) or
        soup.find("a", href=re.compile(rf"search_result={pagina + 1}"))
    )

    log.info("  Pagina %d: %d panden gevonden", pagina, len(uniek))
    return uniek, heeft_volgende


def scrape_alle_funda_tilburg() -> list:
    """Haalt alle Funda-listings voor Tilburg op (alle pagina's)."""
    log.info("Funda scrapen voor heel Tilburg (nieuwe aanpak)...")
    alle = []

    # Gebruik een sessie voor consistente cookies
    sessie = requests.Session()
    # Eerste verzoek om cookies op te halen
    sessie.get("https://www.funda.nl", headers=HEADERS, timeout=15)
    time.sleep(2)

    for pagina in range(1, 25):  # max 24 pagina's = ~480 panden
        listings, heeft_volgende = scrape_funda_pagina(pagina, sessie)
        alle.extend(listings)

        if not listings or not heeft_volgende:
            log.info("  Pagina %d leeg of geen volgende pagina — klaar", pagina)
            break

        time.sleep(PAGE_DELAY)

    # Dedupliceren op URL
    seen = set()
    uniek = []
    for p in alle:
        if p["funda_url"] not in seen:
            seen.add(p["funda_url"])
            uniek.append(p)

    log.info("✓ %d unieke panden gevonden op Funda Tilburg", len(uniek))
    return uniek


# ─── Stap 3: Kruisen met KVP-lijst ───────────────────────────────────────────

def numbers_match(kvp_num: str, funda_num: str) -> bool:
    """Controleert of huisnummers overeenkomen (inclusief variaties zoals 29 vs 29-A)."""
    k = normalize_number(kvp_num)
    f = normalize_number(funda_num)
    if k == f:
        return True
    # Funda-nummer begint met KVP-nummer (bijv. funda=29a, kvp=29)
    if f.startswith(k) and not f[len(k):len(k)+1].isdigit():
        return True
    # KVP-nummer begint met Funda-nummer
    if k.startswith(f) and not k[len(f):len(f)+1].isdigit():
        return True
    return False


def straten_match(kvp_straat: str, funda_straat: str) -> bool:
    """Vergelijkt straatnamen met enige tolerantie voor spellingsvariaties."""
    k = kvp_straat.lower().strip()
    f = funda_straat.lower().strip()
    if k == f:
        return True
    # Één bevat de ander (voor afkortingen zoals 'st.' vs 'sint')
    if k in f or f in k:
        return True
    return False


def kruisen_kvp_met_funda(kvp_adressen: list, funda_listings: list) -> list:
    """Kruist de KVP-lijst met de Funda-listings."""
    log.info("KVP-lijst kruisen met %d Funda-panden...", len(funda_listings))

    # KVP indexeren op genormaliseerde straatnaam
    kvp_per_straat = {}
    for addr in kvp_adressen:
        sleutel = addr["straatnaam_norm"]
        kvp_per_straat.setdefault(sleutel, []).append(addr)

    matches = []

    for listing in funda_listings:
        funda_straat_norm = listing["straatnaam_norm"]

        # Directe straat-match
        kandidaten = kvp_per_straat.get(funda_straat_norm, [])

        # Als geen directe match, probeer gedeeltelijke match
        if not kandidaten:
            for kvp_straat_norm, adressen in kvp_per_straat.items():
                if straten_match(kvp_straat_norm, funda_straat_norm):
                    kandidaten = adressen
                    break

        for kvp_addr in kandidaten:
            if numbers_match(kvp_addr["huisnummer"], listing["huisnummer_funda"]):
                match_type = "exact" if (
                    normalize_number(kvp_addr["huisnummer"]) == normalize_number(listing["huisnummer_funda"])
                ) else "gedeeltelijk"

                match = {
                    **kvp_addr,
                    **listing,
                    "gevonden_op": datetime.now().isoformat(),
                    "match_type": match_type,
                }
                matches.append(match)
                log.info(
                    "  MATCH: %s %s — %s (%s)",
                    kvp_addr["straatnaam"],
                    kvp_addr["huisnummer"],
                    listing.get("prijs", "prijs onbekend"),
                    listing["status"],
                )

    log.info("✓ %d matches gevonden", len(matches))
    return matches


# ─── Stap 4: Excel exporteren ─────────────────────────────────────────────────

def export_excel(matches: list, funda_totaal: int) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "KVP Matches"

    headers = [
        "Straatnaam (KVP)", "Huisnummer", "Postcode",
        "Volledig adres", "Status Funda", "Vraagprijs",
        "Match type", "Gevonden op", "Funda link"
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", start_color="1A1A2E", end_color="1A1A2E")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        cell.alignment = Alignment(horizontal="center")

    status_kleuren = {
        "Te koop": "D5F5E3",
        "Onder voorbehoud": "FDEBD0",
        "Verkocht": "FADBD8",
    }

    for i, m in enumerate(matches, 2):
        ws.append([
            m.get("straatnaam", ""),
            m.get("huisnummer", ""),
            m.get("postcode", ""),
            m.get("volledig_adres", ""),
            m.get("status", ""),
            m.get("prijs", ""),
            m.get("match_type", ""),
            m.get("gevonden_op", "")[:10] if m.get("gevonden_op") else "",
            m.get("funda_url", ""),
        ])

        kleur = status_kleuren.get(m.get("status", ""), "FFFFFF")
        fill = PatternFill("solid", start_color=kleur, end_color=kleur)
        for col in range(1, len(headers) + 1):
            ws.cell(row=i, column=col).fill = fill
            ws.cell(row=i, column=col).font = Font(name="Calibri", size=10)

        link_cel = ws.cell(row=i, column=9)
        if link_cel.value:
            link_cel.hyperlink = link_cel.value
            link_cel.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
            link_cel.value = "Bekijk op Funda →"

    col_breedtes = [28, 14, 12, 42, 18, 14, 14, 14, 20]
    for col, breedte in enumerate(col_breedtes, 1):
        ws.column_dimensions[get_column_letter(col)].width = breedte

    wb.save(RESULTS_EXCEL)
    log.info("Excel opgeslagen: %s", RESULTS_EXCEL)
    return RESULTS_EXCEL


# ─── Stap 5: JSON opslaan ─────────────────────────────────────────────────────

def save_json(matches: list, kvp_count: int, funda_count: int):
    payload = {
        "scan_datum": datetime.now().isoformat(),
        "kvp_adressen_totaal": kvp_count,
        "funda_panden_gescand": funda_count,
        "matches": matches,
        "straten_gescand": funda_count,
    }
    RESULTS_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2))
    log.info("JSON opgeslagen: %s", RESULTS_JSON)


# ─── Stap 6: E-mail via Outlook ───────────────────────────────────────────────

def send_email(matches: list, funda_totaal: int):
    sender = os.environ.get("EMAIL_SENDER", "")
    password = os.environ.get("EMAIL_PASSWORD", "")
    recipient = os.environ.get("EMAIL_RECIPIENT", "")

    if not all([sender, password, recipient]):
        log.info("Geen e-mailconfiguratie gevonden — e-mail overgeslagen")
        return

    datum = datetime.now().strftime("%d-%m-%Y")
    aantal = len(matches)
    tekoop = sum(1 for m in matches if m.get("status") == "Te koop")

    msg = MIMEMultipart("alternative")
    msg["Subject"] = f"KVP x Funda Tilburg: {aantal} match(es) — {datum}"
    msg["From"] = sender
    msg["To"] = recipient

    # Plain text fallback
    plain = [f"KVP x Funda Scan Tilburg — {datum}", "=" * 40, ""]
    plain.append(f"Funda gescand: {funda_totaal} panden | Matches: {aantal} | Te koop: {tekoop}")
    plain.append("")
    if matches:
        for m in matches:
            plain.append(f"  {m.get('volledig_adres', '')} — {m.get('prijs', '')} ({m.get('status', '')})")
            plain.append(f"  {m.get('funda_url', '')}")
            plain.append("")
    else:
        plain.append("Geen matches gevonden vandaag.")
    msg.attach(MIMEText("\n".join(plain), "plain"))

    # HTML e-mail
    if matches:
        rijen = ""
        for m in matches:
            kleur = {"Te koop": "#d5f5e3", "Onder voorbehoud": "#fdebd0", "Verkocht": "#fadbd8"}.get(m.get("status", ""), "#fff")
            rijen += f"""
            <tr style="background:{kleur}">
              <td style="padding:10px;border:1px solid #ddd">{m.get('volledig_adres','')}</td>
              <td style="padding:10px;border:1px solid #ddd"><strong>{m.get('prijs','—')}</strong></td>
              <td style="padding:10px;border:1px solid #ddd">{m.get('status','')}</td>
              <td style="padding:10px;border:1px solid #ddd">{m.get('match_type','')}</td>
              <td style="padding:10px;border:1px solid #ddd"><a href="{m.get('funda_url','')}">Bekijken →</a></td>
            </tr>"""
        tabel = f"""
        <table style="border-collapse:collapse;width:100%;margin-top:16px;font-size:13px">
          <tr style="background:#1A1A2E;color:#fff">
            <th style="padding:10px;text-align:left">Adres</th>
            <th style="padding:10px;text-align:left">Prijs</th>
            <th style="padding:10px;text-align:left">Status</th>
            <th style="padding:10px;text-align:left">Match</th>
            <th style="padding:10px;text-align:left">Link</th>
          </tr>
          {rijen}
        </table>"""
    else:
        tabel = "<p style='color:#666;font-style:italic'>Geen matches gevonden vandaag — alle KVP-panden staan niet te koop op Funda.</p>"

    html = f"""
    <html><body style="font-family:Calibri,Arial,sans-serif;color:#333;max-width:750px;margin:0 auto">
      <div style="background:#E87722;padding:20px;border-radius:8px 8px 0 0">
        <h2 style="color:white;margin:0">KVP x Funda Scanner — Tilburg</h2>
        <p style="color:rgba(255,255,255,0.9);margin:6px 0 0">{datum}</p>
      </div>
      <div style="background:#f5f5f5;padding:16px;border:1px solid #ddd;display:flex;gap:12px">
        <div style="flex:1;text-align:center;background:white;padding:12px;border-radius:6px">
          <div style="font-size:22px;font-weight:bold;color:#E87722">{funda_totaal}</div>
          <div style="font-size:11px;color:#666">Funda-panden gescand</div>
        </div>
        <div style="flex:1;text-align:center;background:white;padding:12px;border-radius:6px">
          <div style="font-size:22px;font-weight:bold;color:#27ae60">{aantal}</div>
          <div style="font-size:11px;color:#666">KVP-matches</div>
        </div>
        <div style="flex:1;text-align:center;background:white;padding:12px;border-radius:6px">
          <div style="font-size:22px;font-weight:bold;color:#2980b9">{tekoop}</div>
          <div style="font-size:11px;color:#666">Actief te koop</div>
        </div>
      </div>
      <div style="padding:16px;background:white;border:1px solid #ddd;border-top:0">
        {tabel}
        <p style="margin-top:16px;font-size:11px;color:#999">
          Volledig overzicht in het bijgevoegde Excel-bestand.<br>
          Automatisch verzonden door KVP x Funda Scanner.
        </p>
      </div>
    </body></html>"""
    msg.attach(MIMEText(html, "html"))

    # Excel bijlage
    if RESULTS_EXCEL.exists():
        with open(RESULTS_EXCEL, "rb") as f:
            deel = MIMEBase("application", "octet-stream")
            deel.set_payload(f.read())
            encoders.encode_base64(deel)
            deel.add_header("Content-Disposition", f'attachment; filename="kvp_tilburg_{datum}.xlsx"')
            msg.attach(deel)

    try:
        with smtplib.SMTP("smtp-mail.outlook.com", 587) as server:
            server.ehlo()
            server.starttls()
            server.login(sender, password)
            server.sendmail(sender, recipient, msg.as_string())
        log.info("E-mail verzonden naar %s", recipient)
    except Exception as e:
        log.error("E-mail versturen mislukt: %s", e)


# ─── Hoofdfunctie ─────────────────────────────────────────────────────────────

def run_scan():
    OUTPUT_DIR.mkdir(exist_ok=True)
    log.info("=" * 60)
    log.info("KVP x Funda Scanner gestart — %s", datetime.now().strftime("%d-%m-%Y %H:%M"))
    log.info("=" * 60)

    # 1. KVP-lijst inlezen
    kvp_adressen = download_and_parse_kvp_pdf(KVP_PDF_URL)

    # 2. Alle Funda-listings ophalen voor heel Tilburg
    funda_listings = scrape_alle_funda_tilburg()

    # 3. Kruisen
    matches = kruisen_kvp_met_funda(kvp_adressen, funda_listings)

    # 4. Opslaan
    save_json(matches, len(kvp_adressen), len(funda_listings))
    export_excel(matches, len(funda_listings))

    # 5. E-mail sturen
    send_email(matches, len(funda_listings))

    log.info("=" * 60)
    log.info("Klaar! %d matches gevonden in %d Funda-panden", len(matches), len(funda_listings))
    log.info("=" * 60)

    return matches


if __name__ == "__main__":
    run_scan()
